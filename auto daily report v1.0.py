import datetime
import openpyxl
from openpyxl.styles import PatternFill
import tkinter

def input_gui(new_wb_name, number_of_columns):
    root = tkinter.Tk()

    def retrieve_input(new_wb_name, number_of_columns):
        outlook_input = text_box.get("1.0", "end-2c")
        if outlook_input == "":
            print("Empty input")
        else:
            create_spreadsheet(outlook_input, new_wb_name, number_of_columns)

    text_box = tkinter.Text(root)
    text_box.pack()
    b_process = tkinter.Button(root, text = "Process", command = lambda: retrieve_input(new_wb_name, number_of_columns))
    b_process.pack()
    root.mainloop()

def create_spreadsheet(outlook_input, new_wb_name, number_of_columns):
    new_wb = openpyxl.Workbook()
    sheet = new_wb["Sheet"]
    
    outlook_input_lines = outlook_input.split("\n")

    row_number = 1
    for line in outlook_input_lines:
        row_list = line.split("\t")
        for i in range(0, number_of_columns):
            cell_ref = sheet.cell(row = row_number, column = i + 1)
            cell_ref.value = row_list[i]
        row_number += 1

    new_wb.save(new_wb_name)

def main():
    print("Which report would you like to process?")
    print("Incorrect Transaction Calculation - 1")
    print("Unknown Card Number - 2")
    print("Sites to Check - 3")
    user_input = input("Process: ")
    if user_input == "1":
        execute_ITC()
    elif user_input == "2":
        execute_UCN()
    elif user_input == "3":
        execute_STC()
    else:
        print("Exiting")

def process_another():
        process_another = input("Process another report? [y/n]: ")
        if process_another == "y":
            main()
        else:
            print("Exiting...")

def get_file_names(report):
    now = datetime.datetime.now()
    if report == "itc":
        new_wb_name =  "D:/Documents/Incorrect Transaction Calculations/incorrect transaction calculations " + str(now.day) + "-" + str(now.month) + "-" + str(now.year) + ".xlsx"
        # prev_wb_name = "D:/Documents/Incorrect Transaction Calculations/incorrect transaction calculations " + str(now.day - 1) + "-" + str(now.month) + "-" + str(now.year) + ".xlsx"
        # when you need to hard code the dates
        prev_wb_name = "D:/Documents/Incorrect Transaction Calculations/incorrect transaction calculations 28-2-2019.xlsx"

    elif report == "ucn":
        new_wb_name =  "D:/Documents/Unknown Card Number/unknown card number " + str(now.day) + "-" + str(now.month) + "-" + str(now.year) + ".xlsx"
        # prev_wb_name = "D:/Documents/Unknown Card Number/unknown card number " + str(now.day - 1) + "-" + str(now.month) + "-" + str(now.year) + ".xlsx"
        # when you need to hard code the dates
        prev_wb_name = "D:/Documents/Unknown Card Number/unknown card number 26-2-2019.xlsx" 

    elif report == "stc":
        new_wb_name =  "D:/Documents/Sites to Check/CompacOnline Status Report_" + "0" + str(now.day) + "0" + str(now.month) + str(now.year) + ".xlsx"
        # prev_wb_name = "D:/Documents/Sites to Check/CompacOnline Status Report_" + str(now.day - 1) + "0" + str(now.month) + str(now.year) + ".xlsx"
        # when you need to hard code the dates
        prev_wb_name = "D:/Documents/Sites to Check/CompacOnline Status Report_28022019.xlsx"

    return prev_wb_name, new_wb_name

def resize(sheet, column_range):
    columns = ["A", "B", "C", "D", "E", "F", "G", "H", "I"]
    rows = sheet.max_row

    for column in columns:
        sheet.column_dimensions[column[:column_range + 1]].width = 30

    for row in range(rows):
        sheet.row_dimensions[row + 1].height = 15

def get_row_lists(prev_sheet, new_sheet):
    prev_row = prev_sheet.iter_rows()
    new_row = new_sheet.iter_rows()

    new_row_list = []
    for row in new_row:
        temp_row = []
        for col in row:
            temp_row += [str(col.value)]
        temp_row += [row[0].row]
        new_row_list += [temp_row]

    prev_row_list = []
    for row in prev_row:
        temp_row = []
        for col in row:
            temp_row += [str(col.value)]
        temp_row += [row[0].fill.start_color.index]
        prev_row_list += [temp_row]

    return prev_row_list, new_row_list

def process_logic(report, reference_column, colour_column, prev_row_list, new_row_list, new_sheet):
    if report == "itc" or report == "ucn":
        for new_row in new_row_list[1:]:
            for prev_row in prev_row_list[1:]:
                if new_row[reference_column] in prev_row and new_row[1] in prev_row:
                    for cell in new_sheet[new_row[colour_column]]:
                        fill_color = PatternFill(fill_type="solid", start_color=prev_row[colour_column], end_color=prev_row[colour_column])
                        cell.fill = fill_color
                    prev_row_list.remove(prev_row)

    elif report == "stc":
        for new_row in new_row_list[1:]:
            for prev_row in prev_row_list[1:]:
                if new_row[3] in prev_row:
                    for cell in new_sheet[new_row[9]]:
                        # copy over "red", "purple", "yellow", "orange", "light blue"
                        if prev_row[9] == "FFFFC000" or prev_row[9] == "FFFF0000" or prev_row[9] == "FF7030A0" or prev_row[9] == "FFFFFF00" or prev_row[9] == "FF00B0F0":
                            fill_color = PatternFill(fill_type="solid", start_color=prev_row[9], end_color=prev_row[9])
                            cell.fill = fill_color
                        # if site hasn't traded in over 10 days mark "light green"
                        elif float(new_row[7]) > 7:
                            fill_color = PatternFill(fill_type="solid", start_color=prev_row[9], end_color=prev_row[9])
                            cell.fill = fill_color
                        # if no card swipe history mark "green"
                        elif new_row[8] == "":
                            fill_color = PatternFill(fill_type="solid", start_color="FF00B050", end_color="FF00B050")
                            cell.fill = fill_color
                        # if card history is older than 2019 mark "green"
                        elif int(new_row[8][new_row[8].find(" ") - 1]) < 9:
                            fill_color = PatternFill(fill_type="solid", start_color="FF00B050", end_color="FF00B050")
                            cell.fill = fill_color
                    prev_row_list.remove(prev_row)

def execute_ITC():
    prev_wb_name, new_wb_name = get_file_names("itc")

    get_outlook = input("Create new spreadsheet? [y/n]: ")
    if get_outlook == "y":
        input_gui(new_wb_name, 7)

    execute = input("Process Incorrect Transaction Calculation? [y/n]: ")
    if execute == "y":
    
        prev_wb = openpyxl.load_workbook(prev_wb_name)
        prev_sheet = prev_wb["Sheet"]
        new_wb = openpyxl.load_workbook(new_wb_name)
        new_sheet = new_wb["Sheet"]

        resize(new_sheet, 6)

        prev_row_list, new_row_list = get_row_lists(prev_sheet, new_sheet)

        process_logic("itc", 3, 7, prev_row_list, new_row_list, new_sheet)

        new_wb.save(new_wb_name)
        print("Done")
        process_another()
    else:
        main()

def execute_UCN():
    prev_wb_name, new_wb_name = get_file_names("ucn")

    get_outlook = input("Create new spreadsheet? [y/n]: ")
    if get_outlook == "y":
        input_gui(new_wb_name, 5)

    execute = input("Process Unknown Card Number? [y/n]: ")
    if execute == "y":
    
        prev_wb = openpyxl.load_workbook(prev_wb_name)
        prev_sheet = prev_wb["Sheet"]
        new_wb = openpyxl.load_workbook(new_wb_name)
        new_sheet = new_wb["Sheet"]

        resize(new_sheet, 4)

        prev_row_list, new_row_list = get_row_lists(prev_sheet, new_sheet)
    
        process_logic("ucn", 4, 5, prev_row_list, new_row_list, new_sheet)

        new_wb.save(new_wb_name)
        print("Done")
        process_another()
    else:
        main()

def execute_STC():
    execute = input("Process Sites to Check? [y/n]: ")
    if execute == "y":
        prev_wb_name, new_wb_name = get_file_names("stc")

        prev_wb = openpyxl.load_workbook(prev_wb_name)
        prev_sheet = prev_wb["Sheet1"]
        new_wb = openpyxl.load_workbook(new_wb_name)
        new_sheet = new_wb["Sheet1"]

        resize(new_sheet, 8)

        prev_row_list, new_row_list = get_row_lists(prev_sheet, new_sheet)
    
        process_logic("stc", 3, 9, prev_row_list, new_row_list, new_sheet)

        new_wb.save(new_wb_name)
        print("Done")
        process_another()
    else:
        main()

main()

# FF00B0F0 = light blue
# FF92D050 = light green
# FFFFC000 = orange
# FF7030A0 = purple
# FF00B050 = green
# FFFF0000 = red
# FFFFFF00 = yellow